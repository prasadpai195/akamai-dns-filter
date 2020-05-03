#! /usr/bin/env python
#Importing the libraries
import requests
import json
from akamai.edgegrid import EdgeGridAuth, EdgeRc
from urllib.parse import urljoin
from openpyxl import load_workbook
import re
import openpyxl, pprint
import sys
import argparse
from datetime import datetime
import calendar
import os
#AUTH SECTION
edgerc_file = os.path.join(os.path.expanduser("~"), '.edgerc')
edgerc = EdgeRc(edgerc_file)
section="dns"
base_url = edgerc.get(section,'host')
baseurl=str('https://')+str(base_url)
client_token=edgerc.get(section,'client_token')
client_secret=edgerc.get(section,'client_secret')
access_token=edgerc.get(section,'access_token')
s = requests.Session()
s.auth = EdgeGridAuth(
client_token=client_token,
client_secret=client_secret,
access_token=access_token
)
if __name__ == '__main__':
    accountSwitchKey="ACCT_SWITCH_KEY";
    wb = load_workbook('input.xlsx')
    sheet = wb['Sheet1']
    sheet2=wb['Sheet2']
    print("Number of entries is:" ,sheet.max_row-1, "hostnames")
    i=2
    for row in range(2, sheet.max_row+1):
        #Reads the hostname  and config namefrom each row.
        zone=sheet['A' + str(row)].value
        print("Running for "+ zone  +"")
        result=s.get(urljoin(baseurl,'/config-dns/v2/zones/'+zone+'/recordsets?sortBy=name%2Ctype&types=CNAME&showAll=false&accountSwitchKey='+ accountSwitchKey +''))
        data=result.json()
        for record in data["recordsets"]:
            sheet2['A' + str(i)].value=zone
            sheet2['B' + str(i)].value=record["name"]
            sheet2['C' + str(i)].value=record["ttl"]
            sheet2['D' + str(i)].value=str(record["rdata"])
            if "edgekey.net" in str(record["rdata"]):
                try:
                    result=s.get('https://'+record["name"]+'/')
                    sheet2['E' + str(i)].value=str(result)
                except:
                    sheet2['E' + str(i)].value="Error"
            wb.save("input.xlsx")
            i+=1;
print("Lets end this program now! For feedback write to prvenkat@akamai.com")
